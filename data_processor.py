import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

# --- 坐标常量定义 ---
DISCOUNT_START_COL = 12   # L列：各银行折扣起始 (L, M, N, O, P)
RESULT_START_COL = 19     # S列：各银行价值起始 (S, T, U, V, W)
HIDDEN_PROD_COL = 26      # Z列：隐藏辅助列

BANKS = ['光大理财', '苏银', '华夏', '联储', '申万']

def get_bank_rate_map(file_source, bank_name):
    if not file_source or not os.path.exists(file_source): return {}
    config = {
        '光大理财': {'k': '证券名称', 'v': '质押率'},
        '苏银': {'k': '证券名称', 'v': '质押率'},
        '华夏': {'k': '债券名称', 'v': '质押率'},
        '联储': {'k': '证券名称', 'v': '折扣'},
        '申万': {'k': '证券名称', 'v': '质押率'}
    }.get(bank_name)
    try:
        df = pd.read_excel(file_source)
        return dict(zip(df[config['k']].astype(str), df[config['v']]))
    except: return {}

def process_excel_logic(template_input, today_input, bank_files_dict):
    # 1. 数据预处理
    df = pd.read_excel(template_input, sheet_name='Sheet1', dtype={'债券代码': str})
    
    # 余额处理 (万元)
    df['余额（元）'] = df['余额（元）'].astype(str).str.replace(',', '').astype(float) / 10000
    df = df[df['余额（元）'] > 0]
    
    # 【核心逻辑】天数 > 60, 行权和到期全部抹除
    df['行权/到期剩余天数'] = pd.to_numeric(df['行权/到期剩余天数'], errors='coerce').fillna(0).astype(int)
    
    # 预清洗日期列为字符串，防止Excel导出格式问题
    df['行权'] = df['行权'].astype(str).replace(['1899-12-31', 'NaT', 'nan'], '')
    df['到期'] = df['到期'].astype(str).replace(['1899-12-31', 'NaT', 'nan'], '')
    
    # 执行抹除
    df.loc[df['行权/到期剩余天数'] > 60, '行权'] = ""
    df.loc[df['行权/到期剩余天数'] > 60, '到期'] = ""
    
    # 估值取整
    df['中债估值'] = pd.to_numeric(df['中债估值'], errors='coerce').fillna(0).astype(int)
    df['中债估值'] = df['中债估值'].clip(upper=100)
    
    # 配置加载
    df_today_cfg = pd.read_excel(today_input, header=None)
    borrow_map = dict(zip(df_today_cfg[0].astype(str), df_today_cfg[1]))
    bank_maps = {name: get_bank_rate_map(bank_files_dict.get(name), name) for name in BANKS}

    # 2. 构造 Excel
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "银行间可用券"
    ws_today = wb.create_sheet("今日")
    ws_sum = wb.create_sheet("汇总")
    bold = Font(bold=True)

    def write_headers(ws):
        headers = ['债券代码', '债券简称', '数量(万元)', '质押率(D)', '金额(E)', 
                   '主体评级', '是否永续', '省份', '估值向下取整', '行权', '到期']
        for i, h in enumerate(headers, 1): ws.cell(1, i, h).font = bold
        for i, b in enumerate(BANKS):
            ws.cell(1, DISCOUNT_START_COL + i, f"{b}折扣").font = bold
            ws.cell(1, RESULT_START_COL + i, b).font = bold

    write_headers(ws_all)
    write_headers(ws_today)

    def fill_rows(ws, product_list, is_today_sheet=False):
        row_idx = 2
        for prod in product_list:
            # 产品标题行：A列显示产品名
            ws.cell(row_idx, 1, prod).font = bold
            if is_today_sheet:
                ws.cell(row_idx, 2, f"借 {borrow_map.get(prod, 0)}w").font = bold
            row_idx += 1
            
            start_row = row_idx
            sub_df = df[df['持有人账户简称'] == prod]
            
            for _, row in sub_df.iterrows():
                ws.cell(row_idx, 1, row['债券代码']) # 明细行A列显示代码
                ws.cell(row_idx, 2, row['债券简称'])
                ws.cell(row_idx, 3, row['余额（元）']) # C列：原始数量
                
                # 质押率默认 0
                ws.cell(row_idx, 4, 0)
                # E = C * D
                ws.cell(row_idx, 5, f"=C{row_idx}*D{row_idx}")
                
                ws.cell(row_idx, 6, row.get('主体评级', ''))
                ws.cell(row_idx, 7, row.get('是否永续', ''))
                ws.cell(row_idx, 8, row.get('省份', ''))
                ws.cell(row_idx, 9, row.get('中债估值', 0))
                ws.cell(row_idx, 10, row['行权'])
                ws.cell(row_idx, 11, row['到期'])
                
                ws.cell(row_idx, HIDDEN_PROD_COL, prod)

                # --- 【修正点】银行计算结果 (S-W) 联动 C 列 (原始数量) ---
                bond_name = str(row['债券简称'])
                for i, bank in enumerate(BANKS):
                    rate = bank_maps[bank].get(bond_name, "")
                    # 写入折扣列 (L-P)
                    ws.cell(row_idx, DISCOUNT_START_COL + i, rate)
                    
                    # 银行计算公式：原始数量(C列) * 银行折扣列
                    # 修正：不再使用E列，而是回归C列
                    discount_col_let = get_column_letter(DISCOUNT_START_COL + i)
                    f = f"=C{row_idx}*{discount_col_let}{row_idx}"
                    ws.cell(row_idx, RESULT_START_COL + i, f)
                row_idx += 1
            
            # 产品汇总行
            end_row = row_idx - 1
            ws.cell(row_idx, 1, "汇总").font = bold
            # 汇总 C(数量), E(金额) 及 S-W(各行结果)
            sum_cols = [3, 5] + list(range(RESULT_START_COL, RESULT_START_COL + len(BANKS)))
            for c in sum_cols:
                col_let = get_column_letter(c)
                ws.cell(row_idx, c, f"=SUM({col_let}{start_row}:{col_let}{end_row})").font = bold
            row_idx += 2

    # 填充明细
    all_prods = df['持有人账户简称'].unique()
    fill_rows(ws_all, all_prods)
    fill_rows(ws_today, [p for p in borrow_map.keys() if p in all_prods], True)
    
    # 3. 汇总页逻辑
    sum_headers = ['产品名字', '今日借', '可用券总计', '调节比例', '最终金额'] + BANKS
    for i, h in enumerate(sum_headers, 1): ws_sum.cell(1, i, h).font = bold

    for r_idx, (prod, target) in enumerate(borrow_map.items(), 2):
        ws_sum.cell(r_idx, 1, prod)
        ws_sum.cell(r_idx, 2, target)
        # 汇总明细页的E列
        ws_sum.cell(r_idx, 3, f"=SUMIF('银行间可用券'!$Z:$Z, $A{r_idx}, '银行间可用券'!$E:$E)")
        ws_sum.cell(r_idx, 4, 1.0)
        ws_sum.cell(r_idx, 5, f"=C{r_idx}*D{r_idx}")
        # 汇总银行结果 (S-W)
        for i in range(len(BANKS)):
            bank_col = get_column_letter(RESULT_START_COL + i)
            ws_sum.cell(r_idx, 6 + i, f"=SUMIF('银行间可用券'!$Z:$Z, $A{r_idx}, '银行间可用券'!{bank_col}:{bank_col})")

    ws_all.column_dimensions['Z'].visible = False
    ws_today.column_dimensions['Z'].visible = False
    return wb

if __name__ == "__main__":
    wb = process_excel_logic('银行间可用券模板.xlsx', 'today.xlsx', {b: f'{b}对券.xlsx' for b in BANKS})
    wb.save("银行间对账_修正版.xlsx")
    print("✅ 修正完毕：银行价值已改为基于 C列 (数量) 计算。")