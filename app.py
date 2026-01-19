import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

# --- 配置 ---
BANKS = ['光大理财', '苏银', '华夏', '联储', '申万']
DISCOUNT_START_COL = 12 
RESULT_START_COL = 19
HIDDEN_PROD_COL = 26 

def get_bank_rate_map(uploaded_file, bank_name):
    config = {
        '光大理财': {'k': '证券名称', 'v': '质押率'},
        '苏银': {'k': '证券名称', 'v': '质押率'},
        '华夏': {'k': '债券名称', 'v': '质押率'},
        '联储': {'k': '证券名称', 'v': '折扣'},
        '申万': {'k': '证券名称', 'v': '质押率'}
    }.get(bank_name)
    try:
        df = pd.read_excel(uploaded_file)
        return dict(zip(df[config['k']].astype(str), df[config['v']]))
    except: return {}

def process_excel_logic(template_file, today_file, bank_files_dict):
    # 1. 数据预处理
    df = pd.read_excel(template_file, sheet_name='Sheet1', dtype={'债券代码': str})
    df['余额（元）'] = df['余额（元）'].astype(str).str.replace(',', '').astype(float) / 10000
    df = df[df['余额（元）'] > 0]
    
    # 日期抹除逻辑
    df['行权/到期剩余天数'] = pd.to_numeric(df['行权/到期剩余天数'], errors='coerce').fillna(0).astype(int)
    df['行权'] = df['行权'].astype(str).replace(['1899-12-31', 'NaT', 'nan', 'None'], '')
    df['到期'] = df['到期'].astype(str).replace(['1899-12-31', 'NaT', 'nan', 'None'], '')
    df.loc[df['行权/到期剩余天数'] > 60, '行权'] = ""
    df.loc[df['行权/到期剩余天数'] > 60, '到期'] = ""
    
    # 估值取整且最高100
    df['中债估值'] = pd.to_numeric(df['中债估值'], errors='coerce').fillna(0).astype(int)
    df['中债估值'] = df['中债估值'].clip(upper=100) 
    
    df_today_cfg = pd.read_excel(today_file, header=None)
    borrow_map = dict(zip(df_today_cfg[0].astype(str), df_today_cfg[1]))
    bank_maps = {name: get_bank_rate_map(bank_files_dict.get(name), name) for name in BANKS}

    # 2. 构造 Excel
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "银行间可用券"
    ws_today = wb.create_sheet("今日")
    ws_sum = wb.create_sheet("汇总")
    bold = Font(bold=True)

    def write_headers(ws):
        headers = ['债券代码', '债券简称', '数量(万元)', '质押率(D)', '金额(E)', 
                   '主体评级', '是否永续', '省份', '估值向下取整', '行权', '到期']
        for i, h in enumerate(headers, 1): ws.cell(1, i, h).font = bold
        for i, b in enumerate(BANKS):
            ws.cell(1, DISCOUNT_START_COL + i, f"{b}折扣").font = bold
            ws.cell(1, RESULT_START_COL + i, b).font = bold

    write_headers(ws_all)
    write_headers(ws_today)

    def fill_rows(ws, product_list, is_today_sheet=False):
        row_idx = 2
        for prod in product_list:
            ws.cell(row_idx, 1, prod).font = bold
            if is_today_sheet:
                ws.cell(row_idx, 2, f"借 {borrow_map.get(prod, 0)}w").font = bold
            row_idx += 1
            start_row = row_idx
            sub_df = df[df['持有人账户简称'] == prod]
            for _, row in sub_df.iterrows():
                ws.cell(row_idx, 1, row['债券代码'])
                ws.cell(row_idx, 2, row['债券简称'])
                ws.cell(row_idx, 3, row['余额（元）'])
                ws.cell(row_idx, 4, 0) # 质押率默认0
                ws.cell(row_idx, 5, f"=C{row_idx}*D{row_idx}") 
                ws.cell(row_idx, 6, row.get('主体评级', ''))
                ws.cell(row_idx, 7, row.get('是否永续', ''))
                ws.cell(row_idx, 8, row.get('省份', ''))
                ws.cell(row_idx, 9, row.get('中债估值', 0))
                ws.cell(row_idx, 10, "" if row['行权'] == "nan" else row['行权'])
                ws.cell(row_idx, 11, "" if row['到期'] == "nan" else row['到期'])
                ws.cell(row_idx, HIDDEN_PROD_COL, prod)
                for i, bank in enumerate(BANKS):
                    rate = bank_maps[bank].get(str(row['债券简称']), "")
                    ws.cell(row_idx, DISCOUNT_START_COL + i, rate)
                    discount_col_let = get_column_letter(DISCOUNT_START_COL + i)
                    f = f"=C{row_idx}*{discount_col_let}{row_idx}" # S-W列联动C列
                    ws.cell(row_idx, RESULT_START_COL + i, f)
                row_idx += 1
            # 组汇总行
            end_row = row_idx - 1
            ws.cell(row_idx, 1, "汇总").font = bold
            sum_cols = [3, 5] + list(range(RESULT_START_COL, RESULT_START_COL + len(BANKS)))
            for c in sum_cols:
                col_let = get_column_letter(c)
                ws.cell(row_idx, c, f"=SUM({col_let}{start_row}:{col_let}{end_row})").font = bold
            row_idx += 2

    fill_rows(ws_all, df['持有人账户简称'].unique())
    fill_rows(ws_today, [p for p in borrow_map.keys() if p in df['持有人账户简称'].unique()], True)
    
    # 汇总页
    sum_headers = ['产品名字', '今日借', '可用券总计', '调节比例', '最终金额'] + BANKS
    for i, h in enumerate(sum_headers, 1): ws_sum.cell(1, i, h).font = bold
    for r_idx, (prod, target) in enumerate(borrow_map.items(), 2):
        ws_sum.cell(r_idx, 1, prod)
        ws_sum.cell(r_idx, 2, target)
        ws_sum.cell(r_idx, 3, f"=SUMIF('银行间可用券'!$Z:$Z, $A{r_idx}, '银行间可用券'!$E:$E)")
        ws_sum.cell(r_idx, 4, 1.0)
        ws_sum.cell(r_idx, 5, f"=C{r_idx}*D{r_idx}")
        for i in range(len(BANKS)):
            bank_col = get_column_letter(RESULT_START_COL + i)
            ws_sum.cell(r_idx, 6 + i, f"=SUMIF('银行间可用券'!$Z:$Z, $A{r_idx}, '银行间可用券'!{bank_col}:{bank_col})")

    ws_all.column_dimensions['Z'].visible = False
    ws_today.column_dimensions['Z'].visible = False
    return wb

# --- Streamlit 界面 ---
st.set_page_config(page_title="质押券自动化整理", layout="centered")
st.title("🏦 质押券自动化整理工具")
st.info("💡 质押率默认0，估值最高100，行权/到期天数>60自动抹除日期。")

# 文件上传区
t_file = st.file_uploader("1. 上传【可用券模板】", type=["xlsx"])
o_file = st.file_uploader("2. 上传【today.xlsx】", type=["xlsx"])
b_files = st.file_uploader("3. 批量上传所有【对券表】", type=["xlsx"], accept_multiple_files=True)

# 自动匹配对券表
matched_banks = {}
if b_files:
    for f in b_files:
        for b in BANKS:
            if b in f.name: matched_banks[b] = f

if st.button("🚀 生成并导出 Excel", use_container_width=True):
    if t_file and o_file:
        try:
            output_wb = process_excel_logic(t_file, o_file, matched_banks)
            output_data = io.BytesIO()
            output_wb.save(output_data)
            output_data.seek(0)
            st.success("处理成功！")
            st.download_button(
                label="📥 点击下载结果文件",
                data=output_data,
                file_name=f"银行间对账_{pd.Timestamp.now().strftime('%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"处理失败: {e}")
    else:
        st.warning("基础底稿缺失，请上传文件1和2。")